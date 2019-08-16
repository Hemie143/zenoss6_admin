import zenAPI.zenApiLib
import re
import time
import argparse
import openpyxl


def get_higher_template(router, device_uid, template_name):
    levels = device_uid.split('/')
    for i in range(len(levels)-2, 3, -1):
        root_uid = '{}/rrdTemplates/{}'.format('/'.join(levels[0:i]), template_name)
        response = router.callMethod('getInfo', uid=root_uid)
        if response['result']['success']:
            # print(root_uid)
            return root_uid
    return

def create_local_template(router, device_uid, template_name):
    # template_uid = get_higher_template(router, device_uid, template_name)
    # if not template_uid:
    #     return
    template_uid = '{}/{}'.format(device_uid, template_name)
    response = router.callMethod('getInfo', uid=template_uid)
    if response['result']['success']:
        # print('Local Template already exists')
        return template_uid
    else:
        response = router.callMethod('makeLocalRRDTemplate', uid=device_uid, templateName=template_name)
        print('***', response)
        return response['result']['tplUid']
    return


def set_datasource_prop(router, template_uid, action, object_name, key, value):
    # Check datasource presence
    # Edit property
    ds_uid = '{}/datasources/{}'.format(template_uid, object_name)
    result = router.callMethod('getInfo', uid=ds_uid)['result']
    if not result['success']:
        print('Datasource not found')
        response = router.callMethod('addDataSource', templateUid=template_uid, name=object_name, type=type)
        print(response)
    else:
        print('Datasource already exists')

    return


object_types = {
    'Datasource': set_datasource_prop
    }


def process_templates(device_router, template_router, data):

    rows = data.max_row
    for r, (template_name, device, component, action) in enumerate(data.iter_rows(min_row=2)):
        result = device_router.callMethod('getDevices', uid=None, params={'name': device.value})['result']
        if result['success'] and result['totalCount'] == 0:
            print('Row {}: Device not found: {}'.format(r + 2, device.value))
            continue
        elif result['success'] and result['totalCount'] > 1:
            print('Row {}: Multiple devices matching name ({})'.format(r + 2, device.value))
            continue
        device_uid = result['devices'][0]['uid']
        # TODO: Check presence of Template ?
        if action.value == 'ADD':
            template_uid = create_local_template(template_router, device_uid, template_name.value)
            # TODO: Check result
            print('Templates Row {}/{}: Template added: {}'.format(r + 2, rows, template_uid))
        elif action.value == 'DELETE':
            print('Templates Row {}/{}: DELETE action to implement'.format(r + 2, rows))
        else:
            print('Templates Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value))
        # route on otype and call
        # f = object_types[otype.value]
        # result = f(tr, template_uid, action.value, oname.value, key.value, value.value)
    return

def process_datasources(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    rows = data.max_row
    for r, (template_name, device, component, action, ds_name, ds_type, key, value) in enumerate(data.iter_rows(min_row=2, max_col=8)):

        # Find device UID
        result = device_router.callMethod('getDevices', uid=None, params={'name': device.value})['result']
        if result['success'] and result['totalCount'] == 0:
            print('Datasources Row {}: Device not found: {}'.format(r + 2, device.value))
            continue
        elif result['success'] and result['totalCount'] > 1:
            print('Datasources Row {}: Multiple devices matching name ({})'.format(r + 2, device.value))
            continue
        device_uid = result['devices'][0]['uid']

        # Find template UID and check presence
        template_uid = '{}/{}'.format(device_uid, template_name.value)
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Datasources Row {}: Template not found'.format(r + 2, template_name))
            continue

        # Find datasource UID and check presence
        ds_uid = '{}/datasources/{}'.format(template_uid, ds_name.value)
        ds_result = template_router.callMethod('getInfo', uid=ds_uid)['result']
        if ds_result['success']:
            # Datasource is present
            if action.value == 'ADD':
                # print('Datasources Row {}: Datasource present'.format(r + 2))
                pass
            elif action.value == 'DELETE':
                result = template_router.callMethod('deleteDataSource', uid=ds_uid)['result']
                if result['success']:
                    print('Datasources Row {}: Datasource deleted'.format(r + 2))
                else:
                    print('Datasources Row {}: ERROR: Datasource delete failed'.format(r + 2))
            else:
                print('Datasources Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))
        else:
            # Datasource is not present
            if action.value == 'ADD':
                response = template_router.callMethod('addDataSource', templateUid=template_uid, name=ds_name.value, type=ds_type.value)
                ds_result = response['result']
                if ds_result['success']:
                    ds_result = template_router.callMethod('getInfo', uid=ds_uid)['result']
                    print('Datasources Row {}: Datasource added'.format(r + 2))
                else:
                    print('Datasources Row {}: Datasource failed to add'.format(r + 2))
                    continue
            elif action.value == 'DELETE':
                print('Datasources Row {}: Datasource already deleted'.format(r + 2))
            else:
                print('Datasources Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))

        # Apply datasource edit
        if key.value:
            if action.value == 'ADD':
                # TODO: don't setInfo if value is already correct
                current_value = ds_result['data'][key.value]
                if isinstance(current_value, unicode):
                    new_value = unicode(str(value.value), 'utf-8')
                else:
                    print('Datasources Row {}: ERROR: Datasource value unknown: {}'.format(r + 2,
                                                                                           type(type(current_value))))
                    print('-- Datasource type current: {}'.format(type(current_value)))
                    print('-- Datasource type new    : {}'.format(type(str(value.value))))
                    print('-- Datasource value current: {}'.format(current_value))
                    print('-- Datasource value new    : {}'.format((str(value.value))))
                    print('-- Datasource test current    : {}'.format((isinstance(current_value, unicode))))
                    continue
                if current_value == new_value:
                    print('Datasources Row {}: Datasource unchanged'.format(r + 2))
                else:
                    data = {u'uid': ds_uid, key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        print('Datasources Row {}: Datasource edited'.format(r + 2))
                    else:
                        print('Datasources Row {}: Datasource failed'.format(r + 2))
            elif action.value == 'DELETE':
                print('Datasources Row {}: DELETE action to implement'.format(r + 2))
            else:
                print('Datasources Row {}: action unknown: {}'.format(r + 2, action.value))
    return

def process_datapoints(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    rows = data.max_row
    for r, (template_name, device, component, action, ds_name, dp_name, key, value) in enumerate(data.iter_rows(min_row=2, max_col=8)):

        # Find device UID
        result = device_router.callMethod('getDevices', uid=None, params={'name': device.value})['result']
        if result['success'] and result['totalCount'] == 0:
            print('Datapoints Row {}: Device not found: {}'.format(r + 2, device.value))
            continue
        elif result['success'] and result['totalCount'] > 1:
            print('Datapoints Row {}: Multiple devices matching name ({})'.format(r + 2, device.value))
            continue
        device_uid = result['devices'][0]['uid']

        # Find template UID and check presence
        template_uid = '{}/{}'.format(device_uid, template_name.value)
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Datapoints Row {}: Template not found'.format(r + 2, template_name))
            continue

        # Find datasource UID and check presence
        ds_uid = '{}/datasources/{}'.format(template_uid, ds_name.value)
        result = template_router.callMethod('getInfo', uid=ds_uid)['result']
        if not result['success']:
            if action.value == 'ADD':
                print('Datapoints Row {}: ERROR: Datasource not found'.format(r + 2, ds_name))
                continue
            elif action.value == 'DELETE':
                print('Datapoints Row {}: ERROR: DELETE action to implement'.format(r + 2))
            else:
                print('Datapoints Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))

        # Create datapoint
        dp_uid = '{}/datapoints/{}'.format(ds_uid, dp_name.value)
        result = template_router.callMethod('getInfo', uid=dp_uid)['result']
        dp_present = result['success']
        # print(result)
        if action.value == 'ADD':
            if dp_present:
                # Datapoint is present
                print('Datapoints Row {}: Datapoint unchanged: {}'.format(r + 2, dp_name.value))
            else:
                # Datapoint doesn't exist yet
                result = template_router.callMethod('addDataPoint', dataSourceUid=ds_uid, name=dp_name.value)['result']
                if result['success']:
                    print('Datapoints Row {}: Datapoint added: {}'.format(r + 2, dp_name.value))
                else:
                    print('Datapoints Row {}: ERROR: Datapoint failed: {}'.format(r + 2, dp_name.value))
        elif action.value == 'DELETE':
            if dp_present and not key.value:
                result = template_router.callMethod('deleteDataPoint', uid=dp_uid)['result']
                if result['success']:
                    print('Datapoints Row {}: Datapoint deleted'.format(r + 2))
                else:
                    print('Datapoints Row {}: ERROR: Datapoint delete failed'.format(r + 2))
                continue
            else:
                print('Datapoints Row {}: Datapoint already deleted'.format(r + 2))
        else:
            print('Datapoints Row {}: action unknown: {}'.format(r + 2, action.value))
            continue

        # Edit datapoint
        if key.value:
            if action.value == 'ADD':
                print('Datapoints Row {}: NOT IMPLEMENTED'.format(r + 2))
                # if result['data'][]
                # TODO: don't setInfo if value is already correct
                '''
                data = {u'uid': ds_uid, key.value: value.value}
                result = template_router.callMethod('setInfo', **data)['result']
                if result['success']:
                    print('Datapoints Row {}: Datasource edited'.format(r + 2))
                else:
                    print('Datapoints Row {}: Datasource failed'.format(r + 2))
                '''
            elif action.value == 'DELETE':
                print('Datapoints Row {}: DELETE action to implement'.format(r + 2))
            else:
                print('Datapoints Row {}: action unknown: {}'.format(r + 2, action.value))

    return

def process_thresholds(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    for r, (template_name, device, component, action, ds_name, ds_type, key, value) in enumerate(data.iter_rows(min_row=2)):

        # Find device UID
        result = device_router.callMethod('getDevices', uid=None, params={'name': device.value})['result']
        if result['success'] and result['totalCount'] == 0:
            print('Datasources Row {}: Device not found: {}'.format(r + 2, device.value))
            continue
        elif result['success'] and result['totalCount'] > 1:
            print('Datasources Row {}: Multiple devices matching name ({})'.format(r + 2, device.value))
            continue
        device_uid = result['devices'][0]['uid']

        # Find template UID and check presence
        template_uid = '{}/{}'.format(device_uid, template_name.value)
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Datasources Row {}: Template not found'.format(r + 2, template_name))
            continue

        # Find datasource UID and check presence
        ds_uid = '{}/datasources/{}'.format(template_uid, ds_name.value)
        result = template_router.callMethod('getInfo', uid=ds_uid)['result']
        if not result['success']:
            if action.value == 'ADD':
                response = router.callMethod('addDataSource', templateUid=template_uid, name=ds_name.value, type=ds_type.value)
                print('Datasources Row {}: Datasource added'.format(r + 2))
                print(response)
            elif action.value == 'DELETE':
                print('Datasources Row {}: DELETE action to implement'.format(r + 2))
            else:
                print('Datasources Row {}: action unknown: {}'.format(r + 2, action.value))

        # Apply datasource edit
        if action.value == 'ADD':
            # TODO: don't setInfo if value is already correct
            data = {u'uid': ds_uid, key.value: value.value}
            result = template_router.callMethod('setInfo', **data)['result']
            if result['success']:
                print('Datasources Row {}: Datasource edited'.format(r + 2))
            else:
                print('Datasources Row {}: Datasource failed'.format(r + 2))
        elif action.value == 'DELETE':
            print('Datasources Row {}: DELETE action to implement'.format(r + 2))
        else:
            print('Datasources Row {}: action unknown: {}'.format(r + 2, action.value))
    return

def process_graphs(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    rows = data.max_row
    for r, (template_name, device, component, action, gr_name, key, value) in enumerate(data.iter_rows(min_row=2, max_col=7)):

        if not device.value:
            continue
        # Find device UID
        result = device_router.callMethod('getDevices', uid=None, params={'name': device.value})['result']
        if result['success'] and result['totalCount'] == 0:
            print('Graphs Row {}: Device not found: {}'.format(r + 2, device.value))
            continue
        elif result['success'] and result['totalCount'] > 1:
            print('Graphs Row {}: Multiple devices matching name ({})'.format(r + 2, device.value))
            continue
        device_uid = result['devices'][0]['uid']

        # Find template UID and check presence
        template_uid = '{}/{}'.format(device_uid, template_name.value)
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Graphs Row {}: Template not found'.format(r + 2, template_name))
            continue

        # Find graph UID and check presence
        gr_uid = '{}/graphDefs/{}'.format(template_uid, gr_name.value)
        gr_result = template_router.callMethod('getInfo', uid=gr_uid)['result']
        gr_present = gr_result['success']
        if not gr_present:
            if action.value == 'ADD':
                response = template_router.callMethod('addGraphDefinition', templateUid=template_uid, graphDefinitionId=gr_name.value)
                gr_result = response['result']
                if gr_result['success']:
                    gr_result = template_router.callMethod('getInfo', uid=gr_uid)['result']
                    print('Graphs Row {}: Graph added'.format(r + 2))
                else:
                    print('Graphs Row {}: ERROR: Graph not added'.format(r + 2))
            elif action.value == 'DELETE':
                print('Graphs Row {}: Graph already deleted'.format(r + 2))
            else:
                print('Graphs Row {}: action unknown: {}'.format(r + 2, action.value))

        # Apply graph edit
        if key.value:
            # print(gr_result)
            current_value = gr_result['data'][key.value]
            if isinstance(current_value, unicode):
                new_value = unicode(str(value.value), 'utf-8')
            elif isinstance(current_value, int):
                new_value = int(value.value)
            else:
                print('Graphs Row {}: ERROR: Graph value unknown: {}'.format(r + 2,type(current_value)))
                print('-- Graphs type current: {}'.format(type(current_value)))
                print('-- Graphs type new    : {}'.format(type(str(value.value))))
                print('-- Graphs value current: {}'.format(current_value))
                print('-- Graphs value new    : {}'.format((str(value.value))))
                print('-- Graphs test current    : {}'.format((isinstance(current_value, unicode))))
                continue
            if action.value == 'ADD':
                if current_value == new_value:
                    print('Graphs Row {}: Graph unchanged'.format(r + 2))
                else:
                    data = {u'uid': gr_uid, key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        print('Graphs Row {}: Graph edited'.format(r + 2))
                    else:
                        print('Graphs Row {}: Graph failed'.format(r + 2))
            elif action.value == 'DELETE':
                print('Graphs Row {}: DELETE action to implement'.format(r + 2))
            else:
                print('Graphs Row {}: action unknown: {}'.format(r + 2, action.value))
    return

def process_graphpoints(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    rows = data.max_row
    for r, (template_name, device, component, action, gr_name, gp_name, dp_name, key, value) in enumerate(data.iter_rows(min_row=2, max_col=9)):

        # Find device UID
        if not device.value:
            continue
        result = device_router.callMethod('getDevices', uid=None, params={'name': device.value})['result']
        if result['success'] and result['totalCount'] == 0:
            print('Graphpoints Row {}: Device not found: {}'.format(r + 2, device.value))
            continue
        elif result['success'] and result['totalCount'] > 1:
            print('Graphpoints Row {}: Multiple devices matching name ({})'.format(r + 2, device.value))
            continue
        device_uid = result['devices'][0]['uid']

        # Find template UID and check presence
        template_uid = '{}/{}'.format(device_uid, template_name.value)
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Graphpoints Row {}: Template not found'.format(r + 2, template_name))
            continue

        # Find graph UID and check presence
        gr_uid = '{}/graphDefs/{}'.format(template_uid, gr_name.value)
        result = template_router.callMethod('getInfo', uid=gr_uid)['result']
        if not result['success']:
            if action.value == 'ADD':
                print('Graphpoints Row {}: ERROR: Graph not found'.format(r + 2, ds_name))
                continue
            elif action.value == 'DELETE':
                print('Graphpoints Row {}: ERROR: DELETE action to implement'.format(r + 2))
            else:
                print('Graphpoints Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))

        # Find datapoint UID and check presence
        # response = template_router.callMethod('getDataPoints', uid=template_router, query='')
        # print(response)
        # continue
        dpp = dp_name.value.split('_')
        ds_name = '_'.join(dpp[:-1])
        dp_name = dpp[-1]
        # /zport/dmd/Devices/Server/Linux/APP/Prod/devices/prb-app-l02.in.credoc.be/Tomcat_JMX/datasources/Auction_ClassLoading_LoadedClassCount/datapoints/LoadedClassCount
        dp_uid = '{}/datasources/{}/datapoints/{}'.format(template_uid, ds_name, dp_name)
        result = template_router.callMethod('getInfo', uid=dp_uid)['result']
        if not result['success']:
            if action.value == 'ADD':
                print('Graphpoints Row {}: ERROR: Datapoint not found'.format(r + 2, ds_name))
                continue
            elif action.value == 'DELETE':
                print('Graphpoints Row {}: ERROR: DELETE action to implement'.format(r + 2))
            else:
                print('Graphpoints Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))

        # Create graphpoint
        # /zport/dmd/Devices/Server/Linux/APP/Prod/devices/prb-app-l02.in.credoc.be/Tomcat_JMX/graphDefs/Auction ClassLoading/graphPoints/ClassLoading_LoadedClassCount_LoadedClassCount
        gp_uid = '{}/graphPoints/{}'.format(gr_uid, gp_name.value)
        result = template_router.callMethod('getInfo', uid=gp_uid)['result']
        gp_present = result['success']
        if action.value == 'ADD':
            if gp_present:
                # Datapoint is present
                print('Graphpoints Row {}: Graphpoint unchanged: {}'.format(r + 2, gp_name.value))
            else:
                # Graphpoint doesn't exist yet
                result = template_router.callMethod('addDataPointToGraph', dataPointUid=dp_uid, graphUid=gr_uid)['result']
                if result['success']:
                    print('Graphpoints Row {}: Graphpoint added: {}'.format(r + 2, gp_name.value))
                else:
                    print('Datapoints Row {}: ERROR: Datapoint failed: {}'.format(r + 2, gp_name.value))
                    print(result)
        elif action.value == 'DELETE':
            if gp_present and not key.value:
                result = template_router.callMethod('deleteDataPoint', uid=gp_uid)['result']
                if result['success']:
                    print('Graphpoints Row {}: Graphpoint deleted'.format(r + 2))
                else:
                    print('Graphpoints Row {}: ERROR: Graphpoint delete failed'.format(r + 2))
                continue
            else:
                print('Graphpoints Row {}: Graphpoint already deleted'.format(r + 2))
        else:
            print('Graphpoints Row {}: action unknown: {}'.format(r + 2, action.value))
            continue

        # Edit datapoint
        if key.value:
            if action.value == 'ADD':
                print('Graphpoints Row {}: NOT IMPLEMENTED'.format(r + 2))
                # if result['data'][]
                # TODO: don't setInfo if value is already correct
                '''
                data = {u'uid': ds_uid, key.value: value.value}
                result = template_router.callMethod('setInfo', **data)['result']
                if result['success']:
                    print('Datapoints Row {}: Datasource edited'.format(r + 2))
                else:
                    print('Datapoints Row {}: Datasource failed'.format(r + 2))
                '''
            elif action.value == 'DELETE':
                print('Graphpoints Row {}: DELETE action to implement'.format(r + 2))
            else:
                print('Graphpoints Row {}: action unknown: {}'.format(r + 2, action.value))

    return


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Manage local templates')
    parser.add_argument('-s', dest='environ', action='store', default='z6_prod')
    parser.add_argument('-f', dest='filename', action='store', default='local_templates.xlsx')
    options = parser.parse_args()
    environ = options.environ
    filename = options.filename

    # Routers
    dr = zenAPI.zenApiLib.zenConnector(section=environ, routerName='DeviceRouter')
    tr = zenAPI.zenApiLib.zenConnector(section=environ, routerName='TemplateRouter')

    # TODO: try
    lt = openpyxl.load_workbook(filename, read_only=True)

    # Templates
    templates_data = lt['Templates']
    process_templates(dr, tr, templates_data)
    exit()
    ds_data = lt['Datasources']
    process_datasources(dr, tr, ds_data)
    dp_data = lt['Datapoints']
    process_datapoints(dr, tr, dp_data)
    gr_data = lt['Graphs']
    process_graphs(dr, tr, gr_data)
    gp_data = lt['Graphpoints']
    process_graphpoints(dr, tr, gp_data)


