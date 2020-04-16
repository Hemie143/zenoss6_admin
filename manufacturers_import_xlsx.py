import zenAPI.zenApiLib
import argparse
import openpyxl


def process_products(router, data):

    result = router.callMethod('getManufacturerList')['result']
    manufacturers = result['data']

    rows = data.max_row
    for r, row_data in enumerate(data.iter_rows(min_row=2, max_col=3)):
        manufacturer, old_productname, new_productname = row_data

        # Check manufacturer
        for m in manufacturers:
            if m['text']['text'] == manufacturer.value:
                break
        else:
            print('Row {}/{}: Manufacturer not found'.format(r+1, rows))
            continue

        # Check old product name
        m_products = router.callMethod('getProductsByManufacturer', uid=m['uid'])['result']['data']
        for p in m_products:
            if p['id'] == old_productname.value:
                break
        else:
            #TODO: check with new name present
            for p in m_products:
                if p['id'] == new_productname.value:
                    print('Row {}/{}: Product already edited.'.format(r+1, rows))
                    break
            else:
                print('Row {}/{}: Product not found'.format(r+1, rows))
            continue
        # print(p)
        p_data = router.callMethod('getProductData', uid=p['uid'], prodname=p['id'])['result']['data']
        if len(p_data) > 1:
            print('p_data is too long: {}'.format(p_data))
        else:
            p_data = p_data[0]
        # print(p_data)
        prodkeys = ",".join(p_data['prodKeys'])
        params = dict(type=p_data['type'],
                      oldname=p_data['name'],
                      prodname=new_productname.value,
                      prodkeys=prodkeys,
                      partno='',
                      description='',
                      uid=p['uid']
                      )
        # print(params)
        result = router.callMethod('editProduct', params=params)['result']
        # print(result)
        if result['success']:
            # TODO: Track job ?
            print('Row {}/{}: Product edited'.format(r+1, rows))
        else:
            print('Row {}/{}: Product edit failed'.format(r+1, rows))




if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Manage Manufacturers')
    parser.add_argument('-s', dest='environ', action='store', default='z6_prod')
    parser.add_argument('-f', dest='filename', action='store', default='manufacturers_import.xlsx')
    options = parser.parse_args()
    environ = options.environ
    filename = options.filename

    # Routers
    mr = zenAPI.zenApiLib.zenConnector(section=environ, routerName='ManufacturersRouter')

    lt = openpyxl.load_workbook(filename, read_only=True)

    products_data = lt['Products']
    process_products(mr, products_data)

