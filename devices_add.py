import zenAPI.zenApiLib
import argparse
import openpyxl
import time

def process_device_sheet(router,job_router, data):
    rows = data.max_row

    state_mapping = {
        'Production': 1000,
        'Staging': 900,
        'Test': 400,
        'Maintenance': 300,
        'Decommissioned': -1,
    }

    for r, (device_class, device, collector, model, state, username) in enumerate(data.iter_rows(min_row=2, max_col=6)):
        if device.value and device_class.value:
            pass
            '''addDevice(self, deviceName, deviceClass, title=None, snmpCommunity="", snmpPort=161, manageIp="", 
            model=False, collector='localhost', rackSlot=0, locationPath="", systemPaths=[], groupPaths=[], 
            productionState=1000, comments="", hwManufacturer="", hwProductName="", osManufacturer="", 
            osProductName="", priority=3, tag="", serialNumber="", zCommandUsername="", zCommandPassword="", 
            zWinUser="", zWinPassword="", zProperties={}, cProperties={})'''
            productionState = state_mapping.get(state.value, 400)
            print('Row {}/{}: Adding device {}'.format(r+1, rows, device.value))
            result = router.callMethod('addDevice',
                                       deviceName=device.value,
                                       deviceClass=device_class.value,
                                       model=False,
                                       productionState=productionState,
                                       zCommandUsername=username.value,
                                       collector=collector.value)['result']
            if result['success']:
                jobid = result['new_jobs'][0]['uuid']
                if track_job(jr, jobid):
                    print('Row {}/{}: Added device {}'.format(r+1, rows, device.value))
                    if model.value in ['Y', True]:
                        # /zport/dmd/Devices/Server/SSH/Linux/devices/acb-webnaban-S01.acc.credoc.be
                        device_uid = '/zport/dmd{}/devices/{}'.format(device_class.value, device.value)
                        result = router.callMethod('remodel', deviceUid=device_uid)['result']
                        jobid = result['jobId']
                        if track_job(jr, jobid):
                            print('Row {}/{}: Job modeling device {} SUCCESS'.format(r+2, rows, device.value))
                        else:
                            print('Row {}/{}: Job modeling device {} FAILED: {}'.format(r+2, rows, device.value, jobid))
                else:
                    print('Row {}/{}: Job adding device {} FAILED: {}'.format(r+2, rows, device.value, jobid))
            else:
                print('Row {}/{}: FAILED to add device {}'.format(r+2, rows, device.value))
        else:
            print('Row {}/{}: Invalid values'.format(r+2, rows))


def track_job(router, jobid, timeout=60):
    job_loop = 0
    timeend = time.time() + timeout
    while True:
        response = router.callMethod('getInfo', jobid=jobid)
        job_status = response['result']['data']['status']
        if job_status == 'SUCCESS':
            return True
        time.sleep(5)
        job_loop += 1
        if time.time() >= timeend or job_loop > 1000:
            return False


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Add device')
    parser.add_argument('-s', dest='section', action='store', default='z6_prod')
    parser.add_argument('-d', dest='device', action='store', default='')
    parser.add_argument('-f', dest='filename', action='store', default='')
    options = parser.parse_args()

    hostname = options.device
    section = options.section
    xlfile = options.filename

    # Routers
    dr = zenAPI.zenApiLib.zenConnector(section=section, routerName='DeviceRouter')
    jr = zenAPI.zenApiLib.zenConnector(section=section, routerName='JobsRouter')

    dt = openpyxl.load_workbook(xlfile, read_only=True)
    devices_data = dt['Devices']
    process_device_sheet(dr, jr, devices_data)
